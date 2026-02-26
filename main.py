"""Automated Timetable Generator - IIIT Dharwad.

Constraint-aware scheduling engine that reads course and room data from Excel
spreadsheets, applies a greedy constraint-satisfaction pipeline, and exports
polished, colour-coded Excel timetables for departments, faculty, and rooms.

Pipeline
--------
1. Load raw Excel sheets for courses and rooms.
2. Normalise each course row into schedulable events.
3. Group events into logical blocks, keeping mutually-shared courses together.
4. Place each block on the timetable grid while respecting rooms, staff, and
   cohort constraints.
5. Export successful placements and report anything unscheduled.
"""

from __future__ import annotations

import random
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill

# ---------------------------------------------------------------------------
# Path configuration
# ---------------------------------------------------------------------------
BASE_DIR: Path = Path(__file__).resolve().parent
INPUT_DIR: Path = BASE_DIR / "inputs"
OUTPUT_DIR: Path = BASE_DIR / "outputs"

# ---------------------------------------------------------------------------
# Grid configuration - five-day week, 15-minute slots
# ---------------------------------------------------------------------------
DAYS: list[str] = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
START_TIME: float = 9.0       # 09:00
END_TIME: float = 18.5        # 18:30
TIME_STEP: float = 0.25       # 15 minutes
BREAK_LENGTH: float = 0.08    # ~5 min buffer between slots
LUNCH_START: float = 12.0
LUNCH_END: float = 14.0
MIN_LUNCH_BREAK: float = 1.0  # minimum free hours within lunch window

# Institute-level reserved time windows (e.g. club hours, meetings).
# Uncomment entries to block off specific slots.
RESERVED_WINDOWS: dict[str, list[tuple[float, float]]] = {
    # "Monday":    [(9.0, 10.0)],
    # "Tuesday":   [(15.0, 16.0)],
    # "Wednesday": [(11.0, 12.0)],
    # "Thursday":  [(14.0, 15.0)],
    # "Friday":    [(10.0, 11.0)],
}

# Colour scheme for Excel output
_CELL_COLOURS: dict[str, PatternFill] = {
    "L":     PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
    "T":     PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    "P":     PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
    "LUNCH": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
}

# Type aliases for readability
Event = dict[str, Any]
Block = dict[str, Any]
Assignment = dict[str, Any]
ScheduleState = dict[str, Any]


# ===================================================================
# Data loading & preprocessing
# ===================================================================

def load_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load and normalise input spreadsheets."""
    try:
        courses_df = pd.read_excel(INPUT_DIR / "courses.xlsx")
        rooms_df = pd.read_excel(INPUT_DIR / "rooms.xlsx")
    except FileNotFoundError as exc:
        print(f"Error: Missing input file. {exc}")
        raise
    except Exception as exc:
        print(f"Error loading data: {exc}")
        raise

    courses_df.columns = courses_df.columns.str.strip().str.lower()
    rooms_df.columns = rooms_df.columns.str.strip().str.lower()
    return courses_df, rooms_df


def preprocess_courses(courses_df: pd.DataFrame) -> list[Event]:
    """Convert course rows into schedulable event dictionaries."""
    events: list[Event] = []
    counter = 1

    for _, row in courses_df.iterrows():
        if str(row.get("schedule or not", "")).strip().lower() != "yes":
            continue

        course_code = str(row["course code"]).strip()
        course_name = str(row["course name"]).strip()
        department = str(row["department"]).strip()
        semester = str(row["semester"]).strip()
        faculty = _nan_safe_str(row.get("faculty", ""))
        lab_assistant = _nan_safe_str(row.get("lab assistant", ""))
        shared_with = _nan_safe_str(row.get("shared with", ""))
        shared_targets = _parse_shared_targets(shared_with)
        lab_type = _nan_safe_str(row.get("lab type", "")).lower()

        basket_raw = str(row.get("basket", "")).strip()
        basket_code = "" if basket_raw.lower() == "nan" else basket_raw.upper()

        num_raw = row.get("number of students", row.get("number or students", 0))
        num_students = int(num_raw if pd.notna(num_raw) else 0)

        lecture_durations = _parse_duration_list(row.get("l"))
        tutorial_durations = _parse_duration_list(row.get("t"))
        practical_hours = _safe_float(row.get("p"))

        for dur in lecture_durations:
            events.append(_build_event(
                counter, "L", dur, course_code, course_name, faculty,
                lab_assistant, department, semester, num_students,
                shared_with, shared_targets, "", basket_code,
            ))
            counter += 1

        for dur in tutorial_durations:
            events.append(_build_event(
                counter, "T", dur, course_code, course_name, faculty,
                lab_assistant, department, semester, num_students,
                shared_with, shared_targets, "", basket_code,
            ))
            counter += 1

        if practical_hours > 0:
            events.append(_build_event(
                counter, "P", practical_hours, course_code, course_name,
                faculty, lab_assistant, department, semester, num_students,
                shared_with, shared_targets, lab_type, basket_code,
            ))
            counter += 1

    return events


# ===================================================================
# Helper utilities - data cleaning & parsing
# ===================================================================

def _nan_safe_str(value: Any) -> str:
    """Convert a value to string, treating NaN / None as empty string."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    result = str(value).strip()
    return "" if result.lower() == "nan" else result


def _safe_float(value: Any) -> float:
    """Return a numeric value as float, defaulting to ``0.0``."""
    if pd.notna(value) and value != "":
        return float(value)
    return 0.0


def _parse_duration_list(value: Any) -> list[float]:
    """Parse comma / semicolon-separated durations into a float list."""
    if value is None or value == "" or (isinstance(value, float) and pd.isna(value)):
        return []

    if isinstance(value, str):
        cleaned = value.replace(";", ",")
        parts = [p.strip() for p in cleaned.split(",") if p.strip()]
        durations: list[float] = []
        for part in parts:
            try:
                dur = float(part)
                if dur > 0:
                    durations.append(dur)
            except ValueError:
                continue
        if durations:
            return durations
        try:
            num = float(value)
            return [num] if num > 0 else []
        except ValueError:
            return []

    try:
        num = float(value)
        return [num] if num > 0 else []
    except (TypeError, ValueError):
        return []


def _build_event(
    counter: int,
    event_type: str,
    duration: float,
    course_code: str,
    course_name: str,
    faculty: str,
    lab_assistant: str,
    department: str,
    semester: str,
    num_students: int,
    shared_with: str,
    shared_targets: list[str],
    lab_type: str,
    basket_code: str,
) -> Event:
    """Create a normalised event dictionary for a single L/T/P occurrence."""
    return {
        "id": f"{course_code}_{event_type}_{counter}",
        "course_code": course_code,
        "course_name": course_name,
        "faculty": faculty,
        "lab_assistant": lab_assistant,
        "department": department,
        "semester": semester,
        "num_students": num_students,
        "type": event_type,
        "duration_hours": duration,
        "lab_type": lab_type,
        "shared_with": shared_with,
        "shared_targets": shared_targets,
        "basket_code": basket_code,
    }


# ===================================================================
# Shared-course detection
# ===================================================================

def _parse_shared_targets(shared_with_value: str) -> list[str]:
    """Normalise 'shared with' text into a list of department tokens."""
    if not shared_with_value:
        return []
    cleaned = shared_with_value.replace(";", ",")
    return [t.strip() for t in cleaned.split(",") if t.strip()]


def _normalize_label(value: str) -> str:
    """Lowercase and strip non-alphanumeric characters for comparison."""
    return "".join(ch for ch in value.lower().strip() if ch.isalnum())


def _normalize_tokens(values: list[str]) -> list[str]:
    """Apply label normalisation to a list of raw tokens."""
    return [_normalize_label(v) for v in values if v.strip()]


def _department_mentions_target(events: list[Event], target_label: str) -> bool:
    """Check if any event's shared targets mention *target_label*."""
    normalised = _normalize_label(target_label)
    if not normalised:
        return False
    return any(
        normalised in set(_normalize_tokens(evt.get("shared_targets", [])))
        for evt in events
    )


def _has_mutual_share(
    events_a: list[Event],
    dept_a: str,
    events_b: list[Event],
    dept_b: str,
) -> bool:
    """Return ``True`` when both departments mutually reference each other."""
    return (
        _department_mentions_target(events_a, dept_b)
        and _department_mentions_target(events_b, dept_a)
    )


def _find_shared_components(dept_events: dict[str, list[Event]]) -> list[list[str]]:
    """Build connected components of departments that mutually share a course."""
    departments = list(dept_events.keys())
    adjacency: dict[str, set[str]] = {d: set() for d in departments}

    for i in range(len(departments)):
        for j in range(i + 1, len(departments)):
            a, b = departments[i], departments[j]
            if _has_mutual_share(dept_events[a], a, dept_events[b], b):
                adjacency[a].add(b)
                adjacency[b].add(a)

    components: list[list[str]] = []
    visited: set[str] = set()
    for dept in departments:
        if dept in visited:
            continue
        stack = [dept]
        component: list[str] = []
        while stack:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            component.append(current)
            stack.extend(adjacency[current] - visited)
        components.append(component)
    return components


# ===================================================================
# Block-building
# ===================================================================

def _next_unassigned_event(
    event_list: list[Event],
    start_index: int,
    assigned: set[str],
) -> tuple[Event | None, int]:
    """Return the next event that has not yet been assigned to a block."""
    idx = start_index
    while idx < len(event_list):
        evt = event_list[idx]
        idx += 1
        if evt["id"] not in assigned:
            return evt, idx
    return None, idx


def build_blocks(events: list[Event]) -> list[Block]:
    """Group events into blocks, honouring cross-scheduled pairs."""
    assigned: set[str] = set()
    blocks: list[Block] = []
    block_counter = 1

    events_by_key: dict[tuple[str, str], list[Event]] = defaultdict(list)
    for evt in events:
        events_by_key[(evt["course_code"], evt["type"])].append(evt)

    for (_code, _etype), course_events in events_by_key.items():
        dept_events: dict[str, list[Event]] = defaultdict(list)
        for evt in course_events:
            dept_events[evt["department"]].append(evt)
        for dept_list in dept_events.values():
            dept_list.sort(key=lambda e: e["id"])

        components = _find_shared_components(dept_events)
        for component in components:
            if len(component) <= 1:
                continue
            pointers = {dept: 0 for dept in component}
            while True:
                block_courses: list[Event] = []
                for dept in component:
                    evt, nxt = _next_unassigned_event(
                        dept_events[dept], pointers[dept], assigned,
                    )
                    if not evt:
                        block_courses = []
                        break
                    block_courses.append(evt)
                    pointers[dept] = nxt
                if not block_courses:
                    break
                for evt in block_courses:
                    assigned.add(evt["id"])
                blocks.append(_build_block(block_counter, block_courses))
                block_counter += 1

    for evt in events:
        if evt["id"] in assigned:
            continue
        blocks.append(_build_block(block_counter, [evt]))
        assigned.add(evt["id"])
        block_counter += 1

    return blocks


def _build_block(block_id: int, courses: list[Event]) -> Block:
    """Aggregate one or more events into a schedulable block entry."""
    first = courses[0]
    total_students = sum(c["num_students"] for c in courses)
    basket_codes = {
        c.get("basket_code", "").strip() for c in courses if c.get("basket_code")
    }
    basket_code = basket_codes.pop() if len(basket_codes) == 1 else ""

    return {
        "id": f"BLOCK_{block_id}",
        "courses": courses,
        "type": first["type"],
        "duration": first["duration_hours"],
        "lab_type": first["lab_type"],
        "total_students": total_students,
        "basket_code": basket_code,
    }


# ===================================================================
# Basket-elective extraction
# ===================================================================

def _course_signature(block: Block) -> tuple:
    """Produce a hashable signature for the course-set in a block."""
    return tuple(sorted(
        (c["course_code"], c["department"], c["semester"])
        for c in block["courses"]
    ))


def extract_basket_blocks(
    blocks: list[Block],
) -> tuple[list[Block], list[Block]]:
    """Split blocks into basket bundles (scheduled first) and remaining."""
    basket_groups: dict[tuple[str, str], list[Block]] = defaultdict(list)
    remaining: list[Block] = []

    for block in blocks:
        code = (block.get("basket_code") or "").strip()
        if not code or block["type"] not in {"L", "T"}:
            remaining.append(block)
            continue
        basket_groups[(code, block["type"])].append(block)

    basket_blocks: list[Block] = []

    for (basket_code, block_type), blocks_list in basket_groups.items():
        blocks_list.sort(key=lambda b: b["id"])
        pointers: dict[tuple, int] = defaultdict(int)
        bundle_counter = 1
        aborted = False

        signatures = {_course_signature(b) for b in blocks_list}

        while True:
            members: list[Block] = []
            durations: set[float] = set()
            lab_types: set[str] = set()

            for sig in sorted(signatures):
                candidates = [
                    b for b in blocks_list if _course_signature(b) == sig
                ]
                idx = pointers[sig]
                if idx >= len(candidates):
                    members = []
                    break
                blk = candidates[idx]
                members.append(blk)
                durations.add(blk["duration"])
                lab_types.add(blk["lab_type"])

            if not members:
                break
            if len(durations) > 1 or len(lab_types) > 1:
                remaining.extend(blocks_list)
                aborted = True
                break

            basket_blocks.append({
                "id": f"BASKET_{basket_code}_{block_type}_{bundle_counter}",
                "basket_code": basket_code,
                "type": block_type,
                "duration": members[0]["duration"],
                "members": members,
            })
            bundle_counter += 1
            for sig in signatures:
                pointers[sig] += 1

        if not aborted:
            for sig in signatures:
                candidates = [
                    b for b in blocks_list if _course_signature(b) == sig
                ]
                while pointers[sig] < len(candidates):
                    remaining.append(candidates[pointers[sig]])
                    pointers[sig] += 1

    return basket_blocks, remaining


def _summarize_basket_block(bundle: Block) -> Block:
    """Flatten a basket bundle into a block-shaped dict for reporting."""
    courses: list[Event] = []
    total_students = 0
    for member in bundle.get("members", []):
        courses.extend(member.get("courses", []))
        total_students += member.get("total_students", 0)
    return {
        "id": bundle.get("id", "UNPLACED_BASKET"),
        "courses": courses,
        "type": bundle.get("type", "L"),
        "duration": bundle.get("duration", 0),
        "lab_type": "",
        "total_students": total_students,
        "basket_code": bundle.get("basket_code", ""),
    }


# ===================================================================
# Room normalisation
# ===================================================================

def _normalize_rooms(rooms_df: pd.DataFrame) -> list[dict[str, Any]]:
    """Normalise room definitions, skipping unusable entries."""
    rooms: list[dict[str, Any]] = []
    for _, row in rooms_df.iterrows():
        code = str(row.get("room code", "")).strip()
        capacity_raw = row.get("room capacity")
        if not code:
            continue
        if pd.isna(capacity_raw) or capacity_raw == "":
            print(f"Skipping room '{code}' due to missing capacity")
            continue
        try:
            capacity = int(float(capacity_raw))
        except (TypeError, ValueError):
            print(f"Skipping room '{code}': invalid capacity '{capacity_raw}'")
            continue
        rooms.append({
            "code": code,
            "capacity": capacity,
            "type": str(row.get("room type", "")).strip().lower(),
        })
    return rooms


# ===================================================================
# Schedule-state management
# ===================================================================

def _init_schedule_state(state: ScheduleState | None = None) -> ScheduleState:
    """Return shared busy maps, creating fresh ones when *state* is ``None``."""
    if state is None:
        return {
            "room_busy": defaultdict(list),
            "faculty_busy": defaultdict(list),
            "lab_assistant_busy": defaultdict(list),
            "group_busy": defaultdict(list),
            "course_day_usage": defaultdict(set),
            "course_slots_by_day": defaultdict(list),
            "faculty_slots_by_day": defaultdict(list),
            "group_slots_by_day": defaultdict(list),
        }
    return state


# ===================================================================
# Constraint helpers
# ===================================================================

def _has_course_spacing(
    course_slots: dict,
    course_key: tuple,
    day: str,
    start: float,
    end: float,
    min_gap: float = TIME_STEP,
) -> bool:
    """Ensure a minimum gap exists between same-course slots on a day."""
    for c_day, c_start, c_end in course_slots.get(course_key, []):
        if c_day != day:
            continue
        if start < c_end + min_gap and start >= c_end:
            return False
        if c_start < end + min_gap and c_start >= end:
            return False
    return True


def _has_lunch_break(
    slots_by_day: dict,
    key: Any,
    day: str,
    start: float,
    end: float,
    min_break: float = MIN_LUNCH_BREAK,
) -> bool:
    """Ensure at least *min_break* free hours within the lunch window."""
    intervals = [
        (c_start, c_end)
        for c_day, c_start, c_end in slots_by_day.get(key, [])
        if c_day == day
    ]
    intervals.append((start, end))

    lunch_intervals = [
        (max(s, LUNCH_START), min(e, LUNCH_END))
        for s, e in intervals
        if e > LUNCH_START and s < LUNCH_END
    ]
    if not lunch_intervals:
        return (LUNCH_END - LUNCH_START) >= min_break

    lunch_intervals.sort()
    merged: list[list[float]] = []
    for s, e in lunch_intervals:
        if not merged or s > merged[-1][1]:
            merged.append([s, e])
        else:
            merged[-1][1] = max(merged[-1][1], e)

    prev_end = LUNCH_START
    for s, e in merged:
        if s - prev_end >= min_break:
            return True
        prev_end = max(prev_end, e)
    return (LUNCH_END - prev_end) >= min_break


def generate_start_times(duration: float) -> list[float]:
    """Produce every feasible start time for a block of *duration* hours."""
    times: list[float] = []
    current = START_TIME
    while current + duration <= END_TIME:
        times.append(current)
        current += TIME_STEP
    return times


def _is_free(
    busy_map: dict,
    key: str,
    day: str,
    start: float,
    end: float,
) -> bool:
    """Return ``True`` if the resource has no overlapping busy slot."""
    for b_day, b_start, b_end in busy_map[key]:
        if b_day == day and not (end <= b_start or b_end <= start):
            return False
    return True


def _mark_busy(
    busy_map: dict,
    key: str,
    day: str,
    start: float,
    end: float,
    include_break: bool = False,
) -> None:
    """Reserve a time window, optionally appending a post-slot buffer."""
    busy_map[key].append((day, start, end))
    if include_break and end < END_TIME:
        busy_map[key].append((day, end, min(end + BREAK_LENGTH, END_TIME)))


def _conflicts_with_lunch(start: float, end: float) -> bool:
    """Lunch overlap is allowed; spacing is enforced elsewhere."""
    return False


def _conflicts_with_reserved_window(
    day: str,
    start: float,
    end: float,
) -> bool:
    """Return ``True`` if the slot intersects any reserved window."""
    for win_start, win_end in RESERVED_WINDOWS.get(day, []):
        if not (end <= win_start or start >= win_end):
            return True
    return False


# ===================================================================
# Room selection
# ===================================================================

def _is_lab(room_type: str) -> bool:
    """Return ``True`` when the room type indicates a lab."""
    return "lab" in (room_type or "")


def _select_single_room(
    block: Block,
    rooms: list[dict],
    day: str,
    start: float,
    end: float,
    room_busy: dict,
    exclude_codes: set[str] | None = None,
) -> list[dict] | None:
    """Pick the tightest-fitting single room for the block.

    Lectures/tutorials are restricted to classrooms; practicals must use
    a lab (optionally matching *lab_type*).
    """
    eligible: list[dict] = []
    for room in rooms:
        if exclude_codes and room["code"] in exclude_codes:
            continue
        if room["capacity"] < block["total_students"]:
            continue

        rtype = room["type"]
        is_lab_room = _is_lab(rtype)

        if block["type"] == "P":
            if not is_lab_room:
                continue
            if block["lab_type"] and rtype != block["lab_type"]:
                continue
        else:
            if is_lab_room:
                continue
            if rtype not in ("classroom", "", None):
                continue

        if not _is_free(room_busy, room["code"], day, start, end):
            continue
        eligible.append(room)

    if not eligible:
        return None
    eligible.sort(key=lambda r: (r["capacity"], r["code"]))
    return [eligible[0]]


def _select_lab_rooms(
    block: Block,
    rooms: list[dict],
    day: str,
    start: float,
    end: float,
    room_busy: dict,
) -> list[dict] | None:
    """Combine multiple labs so their aggregate capacity hosts the block."""
    eligible = [
        r for r in rooms
        if _is_lab(r["type"])
        and (not block["lab_type"] or r["type"] == block["lab_type"])
        and _is_free(room_busy, r["code"], day, start, end)
    ]
    if not eligible:
        return None

    eligible.sort(key=lambda r: r["capacity"], reverse=True)
    selection: list[dict] = []
    remaining = block["total_students"]
    for room in eligible:
        selection.append(room)
        remaining -= room["capacity"]
        if remaining <= 0:
            break
    return None if remaining > 0 else selection


# ===================================================================
# Constraint-check helpers shared by both schedulers
# ===================================================================

def _check_faculty_free(
    faculty_busy: dict,
    lab_assistant_busy: dict,
    courses: list[Event],
    day: str,
    start: float,
    end: float,
) -> bool:
    """Return ``True`` if all faculty *and* lab assistants are free."""
    for course in courses:
        if not _is_free(faculty_busy, course["faculty"], day, start, end):
            return False
        if course["lab_assistant"] and not _is_free(
            lab_assistant_busy, course["lab_assistant"], day, start, end,
        ):
            return False
    return True


def _check_groups_free(
    group_busy: dict,
    courses: list[Event],
    day: str,
    start: float,
    end: float,
) -> bool:
    """Return ``True`` if all department-semester groups are free."""
    return all(
        _is_free(
            group_busy,
            f"{c['department']}_{c['semester']}",
            day, start, end,
        )
        for c in courses
    )


def _check_spacing_and_lunch(
    courses: list[Event],
    day: str,
    start: float,
    end: float,
    course_slots: dict,
    faculty_slots: dict,
    group_slots: dict,
) -> tuple[bool, bool]:
    """Return ``(spacing_ok, lunch_ok)`` for a candidate placement."""
    for course in courses:
        ckey = (
            course["course_code"],
            course["department"],
            course["semester"],
        )
        gkey = f"{course['department']}_{course['semester']}"

        if not _has_course_spacing(course_slots, ckey, day, start, end):
            return False, True

        for slots_map, key in [
            (course_slots, ckey),
            (faculty_slots, course["faculty"]),
            (group_slots, gkey),
        ]:
            if not _has_lunch_break(slots_map, key, day, start, end):
                return True, False

    return True, True


def _record_placement(
    course: Event,
    day: str,
    start: float,
    end: float,
    room_codes: list[str],
    block_type: str,
    basket_code: str,
    state: ScheduleState,
    course_slots_out: list[dict],
) -> None:
    """Update all busy maps and emit a course-slot record for export."""
    _mark_busy(state["faculty_busy"], course["faculty"], day, start, end, True)
    if course["lab_assistant"]:
        _mark_busy(
            state["lab_assistant_busy"], course["lab_assistant"],
            day, start, end, True,
        )
    gkey = f"{course['department']}_{course['semester']}"
    _mark_busy(state["group_busy"], gkey, day, start, end, True)

    ckey = (course["course_code"], course["department"], course["semester"])
    state["course_day_usage"][ckey].add(day)
    state["course_slots_by_day"][ckey].append((day, start, end))
    state["faculty_slots_by_day"][course["faculty"]].append((day, start, end))
    state["group_slots_by_day"][gkey].append((day, start, end))

    course_slots_out.append({
        "course_code": course["course_code"],
        "course_name": course["course_name"],
        "faculty": course["faculty"],
        "department": course["department"],
        "semester": course["semester"],
        "rooms": room_codes,
        "day": day,
        "start": start,
        "end": end,
        "type": block_type,
        "basket_code": basket_code,
    })


# ===================================================================
# Basket-block scheduler
# ===================================================================

def schedule_basket_blocks(
    basket_blocks: list[Block],
    rooms_df: pd.DataFrame,
    state: ScheduleState | None = None,
) -> tuple[list[Assignment], list[dict], list[dict], ScheduleState]:
    """Place basket bundles before the general scheduler runs."""
    if not basket_blocks:
        return [], [], [], _init_schedule_state(state)

    rooms = _normalize_rooms(rooms_df)
    state = _init_schedule_state(state)

    assignments: list[Assignment] = []
    course_slots: list[dict] = []
    unscheduled: list[dict] = []

    blocks_sorted = sorted(
        basket_blocks,
        key=lambda b: (len(b["members"]), b["duration"]),
        reverse=True,
    )

    for bundle_idx, bundle in enumerate(blocks_sorted):
        day_order = (
            DAYS[bundle_idx % len(DAYS):]
            + DAYS[:bundle_idx % len(DAYS)]
        )
        members = bundle["members"]
        start_times = generate_start_times(bundle["duration"])

        placed = False
        day_available = time_available = False
        rooms_issue = faculty_issue = group_issue = False
        spacing_issue = lunch_issue = False

        all_courses = [c for m in members for c in m["courses"]]

        for day in day_order:
            if any(
                day in state["course_day_usage"][
                    (c["course_code"], c["department"], c["semester"])
                ]
                for c in all_courses
            ):
                continue
            day_available = True

            for start in start_times:
                end = start + bundle["duration"]
                if _conflicts_with_lunch(start, end):
                    continue
                if _conflicts_with_reserved_window(day, start, end):
                    continue
                time_available = True

                # Allocate a distinct room per member
                used_rooms: set[str] = set()
                allocations: dict[str, list[dict]] = {}
                alloc_ok = True
                for member in members:
                    alloc = _select_single_room(
                        member, rooms, day, start, end,
                        state["room_busy"], exclude_codes=used_rooms,
                    )
                    if alloc is None:
                        rooms_issue = True
                        alloc_ok = False
                        break
                    codes = [r["code"] for r in alloc]
                    if used_rooms.intersection(codes):
                        rooms_issue = True
                        alloc_ok = False
                        break
                    allocations[member["id"]] = alloc
                    used_rooms.update(codes)
                if not alloc_ok:
                    continue

                if not _check_faculty_free(
                    state["faculty_busy"],
                    state["lab_assistant_busy"],
                    all_courses, day, start, end,
                ):
                    faculty_issue = True
                    continue

                if not _check_groups_free(
                    state["group_busy"], all_courses, day, start, end,
                ):
                    group_issue = True
                    continue

                sp_ok, lu_ok = _check_spacing_and_lunch(
                    all_courses, day, start, end,
                    state["course_slots_by_day"],
                    state["faculty_slots_by_day"],
                    state["group_slots_by_day"],
                )
                if not sp_ok:
                    spacing_issue = True
                    continue
                if not lu_ok:
                    lunch_issue = True
                    continue

                # Commit placement
                for member in members:
                    rcodes = [
                        r["code"] for r in allocations[member["id"]]
                    ]
                    assignments.append({
                        "block_id": member["id"],
                        "day": day,
                        "start": start,
                        "end": end,
                        "rooms": rcodes,
                        "type": member["type"],
                        "courses": member["courses"],
                        "basket_code": bundle.get("basket_code", ""),
                    })
                    for room in allocations[member["id"]]:
                        _mark_busy(
                            state["room_busy"], room["code"],
                            day, start, end, True,
                        )
                    for course in member["courses"]:
                        _record_placement(
                            course, day, start, end, rcodes,
                            member["type"],
                            bundle.get("basket_code", ""),
                            state, course_slots,
                        )

                placed = True
                break
            if placed:
                break

        if not placed:
            reason = _build_failure_reason(
                day_available, time_available,
                rooms_issue, faculty_issue, group_issue,
                spacing_issue, lunch_issue,
                no_day_msg="Basket members already occupy every day",
                no_time_msg="No valid day/time window for basket",
            )
            unscheduled.append({
                "block": _summarize_basket_block(bundle),
                "reason": reason,
            })

    return assignments, course_slots, unscheduled, state


# ===================================================================
# General block scheduler
# ===================================================================

def schedule_blocks(
    blocks: list[Block],
    rooms_df: pd.DataFrame,
    state: ScheduleState | None = None,
    allow_same_day_repeat: bool = False,
) -> tuple[list[Assignment], list[dict], list[dict], ScheduleState]:
    """Assign blocks to day/room slots while tracking busy maps."""
    rooms = _normalize_rooms(rooms_df)
    blocks_sorted = sorted(
        blocks,
        key=lambda b: (
            len(b["courses"]),
            b["total_students"],
            b["duration"],
        ),
        reverse=True,
    )

    state = _init_schedule_state(state)
    assignments: list[Assignment] = []
    course_slots: list[dict] = []
    unscheduled: list[dict] = []

    for block_idx, block in enumerate(blocks_sorted):
        day_order = (
            DAYS[block_idx % len(DAYS):]
            + DAYS[:block_idx % len(DAYS)]
        )
        placed = False
        start_times = generate_start_times(block["duration"])

        day_available = time_available = False
        rooms_issue = faculty_issue = group_issue = False
        spacing_issue = lunch_issue = False

        for day in day_order:
            if not allow_same_day_repeat and any(
                day in state["course_day_usage"][
                    (c["course_code"], c["department"], c["semester"])
                ]
                for c in block["courses"]
            ):
                continue
            day_available = True
            if placed:
                break

            for start in start_times:
                end = start + block["duration"]
                if _conflicts_with_lunch(start, end):
                    continue
                if _conflicts_with_reserved_window(day, start, end):
                    continue
                time_available = True
                if placed:
                    break

                # Room allocation
                allocation = None
                if block["type"] == "P":
                    allocation = _select_single_room(
                        block, rooms, day, start, end, state["room_busy"],
                    )
                    if allocation is None:
                        allocation = _select_lab_rooms(
                            block, rooms, day, start, end,
                            state["room_busy"],
                        )
                else:
                    allocation = _select_single_room(
                        block, rooms, day, start, end, state["room_busy"],
                    )

                if allocation is None:
                    rooms_issue = True
                    continue

                if not _check_faculty_free(
                    state["faculty_busy"],
                    state["lab_assistant_busy"],
                    block["courses"], day, start, end,
                ):
                    faculty_issue = True
                    continue

                if not _check_groups_free(
                    state["group_busy"], block["courses"],
                    day, start, end,
                ):
                    group_issue = True
                    continue

                sp_ok, lu_ok = _check_spacing_and_lunch(
                    block["courses"], day, start, end,
                    state["course_slots_by_day"],
                    state["faculty_slots_by_day"],
                    state["group_slots_by_day"],
                )
                if not sp_ok:
                    spacing_issue = True
                    continue
                if not lu_ok:
                    lunch_issue = True
                    continue

                # Commit placement
                room_codes = [r["code"] for r in allocation]
                assignments.append({
                    "block_id": block["id"],
                    "day": day,
                    "start": start,
                    "end": end,
                    "rooms": room_codes,
                    "type": block["type"],
                    "courses": block["courses"],
                    "basket_code": block.get("basket_code", ""),
                })

                for room in allocation:
                    _mark_busy(
                        state["room_busy"], room["code"],
                        day, start, end, True,
                    )
                for course in block["courses"]:
                    _record_placement(
                        course, day, start, end, room_codes,
                        block["type"],
                        block.get("basket_code", ""),
                        state, course_slots,
                    )

                placed = True
                break

        if not placed:
            reason = _build_failure_reason(
                day_available, time_available,
                rooms_issue, faculty_issue, group_issue,
                spacing_issue, lunch_issue,
                no_day_msg="Course already has a slot on every day",
                no_time_msg=(
                    "No valid day/time window after reserved blocks"
                ),
            )
            unscheduled.append({"block": block, "reason": reason})

    state_out: ScheduleState = {
        "room_busy": state["room_busy"],
        "faculty_busy": state["faculty_busy"],
        "lab_assistant_busy": state["lab_assistant_busy"],
        "group_busy": state["group_busy"],
        "course_day_usage": state["course_day_usage"],
        "course_slots_by_day": state["course_slots_by_day"],
        "faculty_slots_by_day": state["faculty_slots_by_day"],
        "group_slots_by_day": state["group_slots_by_day"],
    }
    return assignments, course_slots, unscheduled, state_out


def _build_failure_reason(
    day_available: bool,
    time_available: bool,
    rooms_issue: bool,
    faculty_issue: bool,
    group_issue: bool,
    spacing_issue: bool,
    lunch_issue: bool,
    no_day_msg: str,
    no_time_msg: str,
) -> str:
    """Produce a human-readable failure reason from diagnostic flags."""
    if not day_available:
        return no_day_msg
    if not time_available:
        return no_time_msg
    conflicts: list[str] = []
    if rooms_issue:
        conflicts.append("no suitable room/lab available")
    if faculty_issue:
        conflicts.append("faculty busy")
    if group_issue:
        conflicts.append("department/semester busy")
    if spacing_issue:
        conflicts.append("requires gap between same-course slots on a day")
    if lunch_issue:
        conflicts.append("requires 1 h free during lunch window")
    return "; ".join(conflicts) if conflicts else "Requirement conflicts"


# ===================================================================
# Retry logic
# ===================================================================

def _retry_unscheduled(
    unscheduled_entries: list[dict],
    rooms_df: pd.DataFrame,
    state: ScheduleState,
    allow_same_day_repeat: bool,
    label: str,
    randomize_order: bool = False,
) -> tuple[list[Assignment], list[dict], list[dict], ScheduleState]:
    """Attempt to place unscheduled blocks again."""
    if not unscheduled_entries:
        return [], [], [], state

    retry_blocks = [entry["block"] for entry in unscheduled_entries]
    if randomize_order:
        random.shuffle(retry_blocks)

    print(
        f"{label}: attempting to place "
        f"{len(retry_blocks)} pending blocks..."
    )
    results = schedule_blocks(
        retry_blocks, rooms_df,
        state=state, allow_same_day_repeat=allow_same_day_repeat,
    )
    retry_assignments, retry_slots, retry_unscheduled, state = results

    if retry_assignments:
        print(f"{label}: recovered {len(retry_assignments)} blocks.")
    if retry_unscheduled:
        print(
            f"{label}: {len(retry_unscheduled)} blocks "
            f"remain unscheduled."
        )
    else:
        print(f"{label}: all pending blocks scheduled.")
    return retry_assignments, retry_slots, retry_unscheduled, state


# ===================================================================
# Excel export
# ===================================================================

def time_to_str(time_float: float) -> str:
    """Format a decimal hour (e.g. 9.5) as ``'09:30'``."""
    hours = int(time_float)
    minutes = int(round((time_float - hours) * 60))
    return f"{hours:02d}:{minutes:02d}"


def get_time_columns() -> list[str]:
    """Generate ordered timetable column labels."""
    columns: list[str] = []
    current = START_TIME
    while current < END_TIME:
        columns.append(time_to_str(current))
        current += TIME_STEP
    return columns


def export_timetables(course_slots: list[dict]) -> None:
    """Create per-department, per-faculty, and per-room Excel timetables."""
    if not course_slots:
        print("No scheduled slots to export.")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    time_cols = get_time_columns()

    def _empty_week() -> dict[str, dict[str, Any]]:
        return {d: {c: None for c in time_cols} for d in DAYS}

    dept_sched: dict[str, dict[str, dict[str, Any]]] = defaultdict(_empty_week)
    faculty_sched: dict[str, dict[str, dict[str, Any]]] = defaultdict(_empty_week)
    room_sched: dict[str, dict[str, dict[str, Any]]] = defaultdict(_empty_week)
    dept_baskets: dict = defaultdict(lambda: defaultdict(set))

    for slot in course_slots:
        room_codes = slot.get("rooms") or (
            [slot["room"]] if slot.get("room") else []
        )
        room_display = ", ".join(room_codes)
        basket_code = slot.get("basket_code", "") or ""
        basket_label = f"BASKET {basket_code}" if basket_code else ""

        course_label = (
            f"{slot['course_code']} | {slot['course_name']}"
            f" | {slot['faculty']} | {room_display}"
        )
        room_label = (
            f"{basket_label} | {course_label}"
            if basket_code
            else course_label
        )
        dept_label = basket_label if basket_code else course_label
        faculty_label = (
            f"{basket_label} | {course_label}"
            if basket_code
            else course_label
        )

        day = slot["day"]
        dept_key = f"{slot['department']}_{slot['semester']}"
        if basket_code:
            bkt_display = (
                f"{slot['course_code']} - {slot['course_name']}"
                f" ({slot['faculty']})"
            )
            dept_baskets[dept_key][basket_code].add(
                (bkt_display, room_display),
            )

        span = max(
            1,
            int(round((slot["end"] - slot["start"]) / TIME_STEP)),
        )
        start_col = time_to_str(slot["start"])

        dept_sched[dept_key][day][start_col] = (
            dept_label, slot["type"], span,
        )
        faculty_sched[slot["faculty"]][day][start_col] = (
            faculty_label, slot["type"], span,
        )
        for rc in room_codes:
            room_sched[rc][day][start_col] = (
                room_label, slot["type"], span,
            )

    _mark_lunch_blocks(dept_sched)
    _mark_lunch_blocks(faculty_sched)
    _mark_lunch_blocks(room_sched)

    _write_department_workbooks(
        dept_sched, dept_baskets, time_cols, "department_timetables",
    )
    _write_individual_workbooks(
        faculty_sched, time_cols, "faculty_timetables",
    )
    _write_individual_workbooks(
        room_sched, time_cols, "room_timetables",
    )
    print(
        "Exported timetables to per-entity Excel files "
        "inside dedicated folders."
    )


def _mark_lunch_blocks(schedule_map: dict) -> None:
    """Fill free lunch-window slots with merged 'LUNCH BREAK' blocks."""
    for schedule in schedule_map.values():
        for day in DAYS:
            current = LUNCH_START
            while current < LUNCH_END:
                col = time_to_str(current)
                if schedule[day].get(col):
                    current += TIME_STEP
                    continue
                start_free = current
                while current < LUNCH_END:
                    if schedule[day].get(time_to_str(current)):
                        break
                    current += TIME_STEP
                end_free = current
                span = max(
                    1,
                    int(round((end_free - start_free) / TIME_STEP)),
                )
                s_col = time_to_str(start_free)
                tmp = start_free
                while tmp < end_free:
                    schedule[day][time_to_str(tmp)] = (
                        "LUNCH BREAK", "LUNCH", 0,
                    )
                    tmp += TIME_STEP
                schedule[day][s_col] = ("LUNCH BREAK", "LUNCH", span)


def _write_timetable_sheet(
    ws: Any,
    schedule: dict,
    time_cols: list[str],
) -> None:
    """Populate a worksheet with timetable rows and merged cells."""
    ws.append(["Day"] + time_cols)
    for day in DAYS:
        row = [day] + [
            schedule[day][col][0] if schedule[day][col] else ""
            for col in time_cols
        ]
        ws.append(row)
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=1).alignment = Alignment(
            wrap_text=True,
        )

        for idx, col in enumerate(time_cols, start=2):
            cell_val = schedule[day][col]
            if not cell_val:
                continue
            cell_obj = ws.cell(row=row_idx, column=idx)
            if isinstance(cell_obj, MergedCell):
                continue
            label, slot_type, span = (
                cell_val
                if len(cell_val) == 3
                else (cell_val[0], cell_val[1], 1)
            )
            cell_obj.value = label
            cell_obj.alignment = Alignment(
                wrap_text=True,
                horizontal="center",
                vertical="center",
            )
            cell_obj.fill = _CELL_COLOURS.get(
                slot_type, PatternFill(),
            )
            if span > 1:
                end_col = min(idx + span - 1, len(time_cols) + 1)
                ws.merge_cells(
                    start_row=row_idx, start_column=idx,
                    end_row=row_idx, end_column=end_col,
                )


def _write_department_workbooks(
    schedule_map: dict,
    basket_map: dict,
    time_cols: list[str],
    folder_name: str,
) -> None:
    """Write department workbooks with timetable + baskets sheet."""
    folder = OUTPUT_DIR / folder_name
    folder.mkdir(parents=True, exist_ok=True)

    for key, schedule in schedule_map.items():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        _write_timetable_sheet(ws, schedule, time_cols)

        bs = wb.create_sheet("Baskets")
        bs.append(["Basket Code", "Course (Faculty)", "Rooms"])
        for bcode, entries in sorted(basket_map.get(key, {}).items()):
            entries_sorted = sorted(entries)
            start_row = bs.max_row + 1
            for i, (cdisplay, rooms) in enumerate(entries_sorted):
                bs.append([
                    bcode if i == 0 else "", cdisplay, rooms,
                ])
                r = bs.max_row
                bs.cell(row=r, column=2).alignment = Alignment(
                    wrap_text=True,
                )
                bs.cell(row=r, column=3).alignment = Alignment(
                    wrap_text=True,
                )
                bs.cell(row=r, column=1).alignment = Alignment(
                    wrap_text=True,
                    horizontal="center",
                    vertical="center",
                )
            if len(entries_sorted) > 1:
                end_row = start_row + len(entries_sorted) - 1
                bs.merge_cells(
                    start_row=start_row, start_column=1,
                    end_row=end_row, end_column=1,
                )
                bs.cell(row=start_row, column=1).alignment = Alignment(
                    wrap_text=True,
                    horizontal="center",
                    vertical="center",
                )

        wb.save(folder / f"{_safe_filename(key)}.xlsx")


def _write_individual_workbooks(
    schedule_map: dict,
    time_cols: list[str],
    folder_name: str,
) -> None:
    """Write per-entity timetables (faculty or room) without basket sheet."""
    folder = OUTPUT_DIR / folder_name
    folder.mkdir(parents=True, exist_ok=True)

    for key, schedule in schedule_map.items():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        _write_timetable_sheet(ws, schedule, time_cols)
        wb.save(folder / f"{_safe_filename(key)}.xlsx")


def _safe_filename(name: str) -> str:
    """Sanitise a string for use as a filename."""
    sanitised = "".join(
        ch if ch.isalnum() or ch in (" ", "_", "-") else "_"
        for ch in name
    )
    return sanitised.strip().replace(" ", "_") or "timetable"


def export_unscheduled(entries: list[dict]) -> None:
    """Write a spreadsheet summarising every unscheduled block."""
    if not entries:
        print("All blocks scheduled successfully.")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows: list[dict] = []
    for entry in entries:
        block = entry["block"]
        reason = entry.get("reason", "")
        details: list[str] = []
        faculties: list[str] = []
        departments: set[str] = set()
        semesters: set[str] = set()
        for course in block["courses"]:
            details.append(
                f"{course['course_code']} ({course['course_name']})",
            )
            faculties.append(course["faculty"])
            departments.add(course["department"])
            semesters.add(str(course["semester"]))
        rows.append({
            "block_id": block["id"],
            "type": block["type"],
            "duration_hours": block["duration"],
            "total_students": block["total_students"],
            "lab_type": block["lab_type"],
            "courses": "; ".join(details),
            "faculties": "; ".join(faculties),
            "departments": "; ".join(sorted(departments)),
            "semesters": "; ".join(sorted(semesters)),
            "reason": reason,
        })

    output_file = OUTPUT_DIR / "unscheduled_blocks.xlsx"
    pd.DataFrame(rows).to_excel(output_file, index=False)
    print(f"Saved {len(entries)} unscheduled blocks to {output_file}")


# ===================================================================
# Main entry point
# ===================================================================

def main() -> None:
    """Orchestrate the end-to-end scheduling pipeline."""
    try:
        print("Loading data...")
        courses_df, rooms_df = load_data()

        print("Preprocessing courses...")
        events = preprocess_courses(courses_df)
        print(f"Generated {len(events)} event blocks.")

        print("Building logical blocks...")
        blocks = build_blocks(events)
        print(f"Prepared {len(blocks)} blocks for scheduling.")

        basket_blocks, remaining_blocks = extract_basket_blocks(blocks)
        print(
            f"Detected {len(basket_blocks)} basket bundles "
            f"and {len(remaining_blocks)} standard blocks."
        )

        assignments: list[Assignment] = []
        course_slots: list[dict] = []
        final_unscheduled: list[dict] = []

        bkt_assign, bkt_slots, bkt_unsched, state = (
            schedule_basket_blocks(basket_blocks, rooms_df)
        )
        assignments.extend(bkt_assign)
        course_slots.extend(bkt_slots)
        final_unscheduled.extend(bkt_unsched)

        print("Scheduling remaining blocks with break buffers...")
        reg_assign, reg_slots, unscheduled, state = schedule_blocks(
            remaining_blocks, rooms_df, state=state,
        )
        assignments.extend(reg_assign)
        course_slots.extend(reg_slots)

        if unscheduled:
            count = len(unscheduled)
            print(
                f"Initial pass left {count} blocks unscheduled. "
                f"Retrying..."
            )
            for attempt in range(3):
                if not unscheduled:
                    break
                label = f"Retry pass #{attempt + 1}"
                r_assign, r_slots, unscheduled, state = (
                    _retry_unscheduled(
                        unscheduled, rooms_df, state,
                        allow_same_day_repeat=True,
                        label=label,
                        randomize_order=(attempt == 2),
                    )
                )
                assignments.extend(r_assign)
                course_slots.extend(r_slots)
                if not r_assign and unscheduled:
                    print(
                        f"{label}: no additional placements possible.",
                    )

        final_unscheduled.extend(unscheduled)

        print("Exporting scheduled timetables...")
        export_timetables(course_slots)

        if final_unscheduled:
            export_unscheduled(final_unscheduled)
        else:
            print(
                "All blocks scheduled. "
                "No unscheduled blocks to report."
            )

        print("Done.")
    except Exception as exc:
        print(f"An error occurred: {exc}")
        raise


if __name__ == "__main__":
    main()
